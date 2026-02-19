import os
from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.implementations.azure_blob_storage import AzureBlobStorage
from app.interfaces.excel_interface import IExcelRepository

# blob name where workbook will be stored; environment variable only
EXCEL_BLOB_NAME = os.environ.get('EXCEL_BLOB_NAME', 'transformed_data/output.xlsx')


class ExcelRepository(IExcelRepository):
    def __init__(self):
        self.storage = AzureBlobStorage()
        self.blob_name = EXCEL_BLOB_NAME
        
        # Try to fetch existing workbook from blob, or create new one
        try:
            data = self.storage.get(self.blob_name)
            self.wb = load_workbook(BytesIO(data))
        except Exception:
            # Create new workbook if it doesn't exist in blob
            self.wb = Workbook()
            ws = self.wb.active
            ws.append(['id', 'filename', 'file_type', 'json_data'])

    def _sync_to_blob(self):
        """Save workbook to blob storage"""
        bio = BytesIO()
        self.wb.save(bio)
        bio.seek(0)
        self.storage.save(self.blob_name, bio.read())

    def append(self, record: dict) -> int:
        ws = self.wb.active
        row = ws.max_row + 1
        ws.append([
            record.get('id'), 
            record.get('filename'),
            record.get('file_type'),
            record.get('json_data')
        ])
        # Save to blob storage
        self._sync_to_blob()
        return row

    def get_stream(self) -> BytesIO:
        # Fetch fresh copy from blob storage
        try:
            data = self.storage.get(self.blob_name)
            bio = BytesIO(data)
            return bio
        except Exception:
            raise FileNotFoundError(f"Excel file {self.blob_name} not found in blob storage")
