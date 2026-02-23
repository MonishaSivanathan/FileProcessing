import json
from openpyxl import load_workbook
from flask import jsonify, request

from app.logging_config import get_logger
from app.services.document_service import DocumentService
from app.utils.file_validator import validate_file

service = DocumentService()
logger = get_logger(__name__)


def upload_document():
    logger.info("Upload request received")
    file = request.files.get("file")
    if not file:
        logger.warning("Upload request rejected: missing file part 'file'")
        return jsonify({"error": "No file provided"}), 400

    try:
        validate_file(file)
    except ValueError as e:
        logger.warning("Upload request rejected: validation failed (%s)", str(e))
        return jsonify({"error": str(e)}), 400

    try:
        result = service.process_upload(file.stream, file.filename)
    except ValueError as e:
        logger.warning("Upload request rejected by service: %s", str(e))
        return jsonify({"error": str(e)}), 400
    except Exception:
        logger.exception("Upload request failed due to an unexpected server error")
        return jsonify({"error": "Internal server error"}), 500
    logger.info(
        "Upload request completed: filename='%s', document_id='%s'",
        file.filename,
        result.id,
    )
    return jsonify(result.__dict__), 201


def get_excel():
    try:
        logger.info("Excel fetch request received")
        stream = service.get_excel_stream()
        # parse workbook in-memory and return structured JSON for all entries

        stream.seek(0)
        wb = load_workbook(stream)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return jsonify({"data": []}), 200

        headers = [h for h in rows[0]]
        # common column names that may contain transformed JSON
        json_cols = [
            name for name in ("json_data", "transformed_data") if name in headers
        ]
        json_col_idx = headers.index(json_cols[0]) if json_cols else None

        # locate id/filename/file_type indices if present
        id_idx = headers.index("id") if "id" in headers else None
        filename_idx = headers.index("filename") if "filename" in headers else None
        filetype_idx = headers.index("file_type") if "file_type" in headers else None

        entries = []
        for r in rows[1:]:
            entry = {
                "id": r[id_idx] if id_idx is not None else None,
                "filename": r[filename_idx] if filename_idx is not None else None,
                "file_type": r[filetype_idx] if filetype_idx is not None else None,
                "transformed_data": [],
            }

            if json_col_idx is not None:
                cell = r[json_col_idx]
                if not cell:
                    entry["transformed_data"] = []
                elif isinstance(cell, str):
                    lines = [l for l in cell.splitlines() if l.strip()]
                    parsed = []
                    for ln in lines:
                        try:
                            val = json.loads(ln)
                        except Exception:
                            # if the cell is a JSON array
                            try:
                                val = json.loads(cell)
                                if isinstance(val, list):
                                    parsed.extend(val)
                                    break
                            except Exception:
                                val = ln
                        if isinstance(val, list):
                            parsed.extend(val)
                        else:
                            parsed.append(val)
                    entry["transformed_data"] = parsed
                else:
                    # non-string (unlikely) - include raw
                    entry["transformed_data"] = [cell]
            else:
                # reconstruct object from all columns except id/filename/file_type
                obj = {}
                for k, val in zip(headers, r):
                    if k in ("id", "filename", "file_type"):
                        continue
                    obj[k] = val
                entry["transformed_data"] = [obj]

            entries.append(entry)

        logger.info(
            "Excel fetch request completed successfully: returned_entries=%d",
            len(entries),
        )
        return jsonify({"data": entries}), 200
    except FileNotFoundError:
        logger.warning("Excel fetch request failed: workbook file not found")
        return jsonify({"error": "Excel file not found"}), 404
    except Exception:
        logger.exception("Excel fetch request failed due to an unexpected server error")
        return jsonify({"error": "Internal server error"}), 500
